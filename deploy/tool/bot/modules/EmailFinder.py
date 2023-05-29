import requests
# import undetected_chromedriver as uc
from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
import pickle
import os
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import datetime
import webdriver_manager
import pandas as pd
import pyautogui
import clipboard
import csv
from openpyxl import load_workbook
from openpyxl import load_workbook,Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import random
from modules.sharePoint_api import API
class Tool:
    def __int__(self):
        self.sheet_name=None
        self.workbook_name=None
    def create_writer(self):
        self.writer = API()
        self.writer.authenticate()
        print("writer created")
    def initialize(self,email,password,user_dir=True):
        with open("settings.txt","r") as file:
            lines=file.readlines()
            user_dir_data=lines[-1].replace("\n","")
            user_profile=user_dir_data.split("\\")[-1]
            user_dir_data=user_dir_data.replace("\\"+user_profile,"")
        # Initiate the ChromeDriver
        options = webdriver.ChromeOptions()
        options.add_argument('--disable-infobars')
        # options.add_argument('--disable-extensions')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-setuid-sandbox')
        options.add_argument('--disable-gpu')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-logging')
        options.add_argument('--disable-browser-side-navigation')
        options.add_argument('--remote-debugging-address=0.0.0.0')
        options.add_argument('--remote-debugging-port=9222')
        options.add_argument('--remote-debugging-port=0')
        options.add_argument('--disable-features=VizDisplayCompositor')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_argument('--start-maximized')
        options.add_argument('--disable-popup-blocking')
        options.add_argument('--disable-notifications')

        if user_dir==True:
            options.add_argument(f'--user-data-dir={user_dir_data}')
            options.add_argument(f'--profile-directory={user_profile}')

        # initiate the driver
        # driver = uc.Chrome()
        self.driver = webdriver.Chrome(chrome_options=options)
        self.driver.get("https://www.linkedin.com")
        sleep(0.7)

        try:
            cookies = pickle.load(open("cookies.pkl", "rb"))
            for cookie in cookies:
                self.driver.add_cookie(cookie)
            self.driver.get("https://www.linkedin.com/")
        except:
            self.driver.get("https://www.linkedin.com/login")

            # wait for the page to load
            sleep(8)

            # find email and password fields and fill them in
            email_field = self.driver.find_element(by='xpath',value="//input[@id='username']")
            email_field.send_keys(email)

            password_field = self.driver.find_element(by='xpath',value="//input[@id='password']")
            password_field.send_keys(password)

            # click the login button
            login_button = self.driver.find_element(by='xpath',value="//button[@type='submit']")
            login_button.click()

            # wait for the login to complete
            sleep(5)
            input("Hit Enter After Verification")

            # check if the login was successful
            if self.driver.current_url == "https://www.linkedin.com/feed/":
                print("Login successful!")
                with open("cookies.pkl", "wb") as f:
                    pickle.dump(self.driver.get_cookies(), f)

            else:
                print("Login failed. Please check your email and password.")
    def randomize_sleep(self,long=False):
        if long==False:
            mini=random.randrange(4,6)
            maxi=random.randrange(8,10)
            random_sleep=random.randrange(mini,maxi)
            sleep(random_sleep)
        else:
            mini = random.randrange(18, 21)
            maxi = random.randrange(27, 30)
            random_sleep = random.randrange(mini, maxi)
            sleep(random_sleep)
    def dict_to_list(self,data):
        temp=[]
        for d in data:
            temp.append(d['Linkedin'])
        return temp
    def get_linkedIn_data(self):
        linkedin_data,existing_data=self.writer.read_list_data()
        existing_data = self.dict_to_list(existing_data)
        return linkedin_data,existing_data
    def check_url(self,link):
        self.driver.get(link)
        sleep(5)
        if self.driver.current_url=="https://www.linkedin.com/404/":
            return False
        else:
            return True
    def update_number_email(self,file_path, linkedin_value, new_number, new_email):
        # Read the "Sales Navigator" subsheet into a pandas dataframe
        df = pd.read_excel(file_path, sheet_name="SalesNavigator")

        # Find the index of the row where the specified linkedin value exists
        index = df.index[df['Linkedin'] == linkedin_value].tolist()

        # If the linkedin value does not exist in the dataframe, return an error message
        if not index:
            return "Error: Linkedin value not found in dataframe"

        # Update the "Number" and "Email" columns for the row with the specified index
        df.loc[index, "Phone Number"] = new_number
        df.loc[index, "Email"] = new_email

        # Write the updated dataframe back to the Excel file
        df.to_excel(file_path, sheet_name="SalesNavigator", index=False)

        return "Number and Email columns updated successfully"
    def scrape_apollo(self,link):
        # self.driver.get(link)
        # sleep(10)

        btn = self.find_until("xpath","//span[text()='View email address']//parent::div",delay=30)
        if btn!=None:
            action = ActionChains(self.driver)
            action.move_to_element(btn).click().perform()
            try:
                action = ActionChains(self.driver)
                action.move_to_element(btn).click().perform()
            except:
                pass
            # btn.click()
            sleep(1)
        email= self.find_until("xpath","//div[text()='No email']",delay=90)
        if email==None:
            print("Found email")
            email=self.find_until('xpath',"//div[text()='Business']/parent::div",delay=160)
            if email!=None:
                email=email.text.split("\n")[0]
                print(email)
                try:
                    domain= email.split("@")[1]
                    print(domain)
                except:
                    pass
        else:
            email=None

        number=self.find_until("xpath","//div[text()='HQ']/parent::div",delay=30)
        if number!=None:
            number=number.text.split("\n")[0]
            if "No phone number found" in number or "ViewHQnumber" in number:
                number=None
            print(number)
        if email=="Verifying" or email=="View email address":
            email=None
        return email,number
        # self.update_number_email("data.xlsx",link,number,email)
    def scrape_salesql(self,link):
        # self.driver.get(link)
        # sleep(10)
        image_location = pyautogui.locateOnScreen('salesql.png')
        image_center = pyautogui.center(image_location)
        pyautogui.click(image_center)
        for i in range(30):
            try:
                image_location = pyautogui.locateOnScreen('add.png')
                image_center = pyautogui.center(image_location)
                pyautogui.click(image_center)
                break
            except:
                pass
        sleep(3)
        for i in range(15):
            print("finding see plans")
            try:
                image_location = pyautogui.locateOnScreen('see_plans.png')
                image_center = pyautogui.center(image_location)
                # pyautogui.click(image_center)
                # for i in range(200):
                #     try:
                #         image_location = pyautogui.locateOnScreen('close.png')
                #         image_center = pyautogui.center(image_location)
                #         pyautogui.click(image_center)
                #         pyautogui.moveTo(100, 100, duration=0.1)
                #         break
                #     except:
                #         pass
                print(image_center)
                if image_center!=None or image_center!="":
                    return None
                break
            except:
                pass
        for i in range(50):
            try:
                image_location = pyautogui.locateOnScreen('copy.png')
                image_center = pyautogui.center(image_location)
                pyautogui.click(image_center)
                # sleep(3)
                email = clipboard.paste()
                if "@" in email:
                    pass
                else:
                    email=None
                break
            except Exception as e:

                email=None
                pass
        try:
            if email=="":
                email=None
            print(email)
        except:
            email=None
        # Send the Escape key to the browser window
        for i in range(50):
            try:
                image_location = pyautogui.locateOnScreen('close.png')
                image_center = pyautogui.center(image_location)
                pyautogui.click(image_center)
                pyautogui.moveTo(100, 100, duration=0.1)
                break
            except:
                pass
        return email
        # action=ActionChains(self.driver)
        # btn=self.find_until("xpath","//div[@class='action-button success pointer']")
        # action.move_to_element(btn).click().perform()
    def scrape_pipileads(self,link):
        # self.driver.get(link)
        # sleep(8)
        link=self.driver.current_url
        if "%" in link:
            return None
        self.driver.get('https://pipileads.com/app/social-search/')
        field=self.find_until("xpath","//input[@placeholder='Profile URL of Linkedin, Facebook, Twitter, Github, Etc']")
        field.send_keys(link)
        btn=self.find_until("xpath","//div[@class='drop-down-list-div']//button")
        actions=ActionChains(self.driver)
        actions.move_to_element(btn).click().perform()
        sleep(13)
        self.driver.get('https://pipileads.com/app/social-search/')
        email=None
        emails=self.find_until("xpath","//tbody//tr[1]//span[@class='person-email']",multiple=True)
        # print(len(emails))
        if emails!=None:
            for e in emails:

                email=e.text+","
            email=email.replace(" ","")
        print(email)
        if email=="" or email=="," or email=='Searching...,':
            email=None
        return email
    def scrape_kendo(self,link):
        email=self.find_until("xpath","//p[@class='max-w-[18rem] truncate m-0 inline-block align-sub ']")
        if email!=None:
            email=email.text
            if email=="No data found for now" or "*" in email:
                email=None
        return email

    def find_until(self,selector,value,multiple=False,delay=700):
        for i in range(delay):
            try:
                if multiple==True:
                    elem = self.driver.find_elements(by=selector, value=value)
                    return elem
                elem=self.driver.find_element(by=selector,value=value)
                return elem
            except:
                pass
        return None
    def scrape(self,profile):
        self.driver.get(profile)
        name = self.find_until('xpath',"//div[@class='mt2 relative']//h1")
        if name!=None:
            name=name.text
        # company = self.find_until('xpath',"//a[@data-field='experience_company_logo'][1]/parent::div[1]/following-sibling::div//span[@class='t-14 t-normal']")
        # if company!=None:
        #     company=company.text.split("\n")[0]
        location = self.find_until('xpath',
                                  "//a[@data-field='experience_company_logo'][1]/parent::div[1]/following-sibling::div//span[@class='t-14 t-normal t-black--light'][2]")
        if location != None:
            location = location.text.split("\n")[0]
        job_titles = self.find_until("xpath","//section[@tabindex='-1']//span[@class='mr1 t-bold']",multiple=True)
        jt=0
        if job_titles!=None:
            for jt,job_title in enumerate(job_titles):
                job_title=job_title.text.split("\n")[0]
                if "founder" in job_title.casefold():
                    break
        jt=str(jt+1)
        company_link = self.find_until("xpath",f"(//a[@data-field='experience_company_logo'])[{jt}]")
        if company_link!=None:
            company_link=company_link.get_attribute("href")

            self.driver.get(company_link+"about")
            company_name = self.find_until("xpath","//h1")
            if company_name!=None:
                company_name=company_name.text
            website=self.find_until('xpath',"//dt[text()='Website']/following-sibling::dd[1]/a")
            if website!=None:
                website=website.get_attribute("href")
            try:
                phone = self.driver.find_element(by='xpath',value="//dt[text()='Phone']/following-sibling::dd/a/span").text
            except:
                phone = None
            employee=self.find_until('xpath',"//dt[text()='Company size']/following-sibling::dd[1]")
            if employee!=None:
                employee=employee.text

        # print(name,company,company_link,location,job_title,website,phone,employee)
        print([name],"name")
        print([name.split(" ")[0],"==First Name"])
        print([name.split(" ")[1], "==Last Name"])
        print([company_name],"==company")
        print([company_link],"==link")
        print([location],"==location")
        print([job_title],"==title")
        print([website],"==website")
        print([phone],"==phone")
        print([employee],"==employee")
    def read_history(self):
        try:
            with open("history.txt","r") as file:
                lines=file.readlines()
                for i in range(len(lines)):
                    lines[i] = lines[i].replace("\n","")
            return lines
        except:
            with open("history.txt","w") as file:
                pass
            lines=[]
            return lines
    def write_history(self,data):
        try:
            with open("history.txt","a+",encoding='utf-8') as file:
                file.writelines(data['Linkedin']+"\n")
        except:
            with open("history.txt","w",encoding='utf-8') as file:
                file.writelines(data['Linkedin']+"\n")
    def load_book_main(self):
        try:
            # Load the existing workbook
            self.wb = load_workbook(self.workbook_name)
            # Check if the sheet already exists
            if self.sheet_name in self.wb.sheetnames:
                # If the sheet exists, select it
                self.ws = self.wb[self.sheet_name]
            else:
                # If the sheet doesn't exist, create a new one
                self.ws = self.wb.create_sheet(self.sheet_name)

        except Exception as e:
            self.wb=Workbook()
            self.ws = self.wb.active
            self.ws.title = self.sheet_name
    def save_book_main(self):
        # Save the workbook
        self.wb.save(self.workbook_name)
    def write_data_v2(self, data):
        self.writer.update_list_data(data)
        self.write_history(data)


    def write_data(self,data):
        try:
            # Load the existing workbook
            wb = load_workbook(self.workbook_name.replace(".xlsx","")+'ScrapedData.xlsx')

            # Select the active worksheet
            ws = wb.active

            # Add some data rows

            ws.append(data)

            # Save the workbook
            wb.save('ScrapedData.xlsx')
        except:
            with open("ScrapedData.xlsx", "w", encoding='utf-8', newline="") as file:
                pass

        # try:
        #     with open("ScrapedData.csv","a",encoding='utf-8',newline="") as file:
        #         writer = csv.writer(file)
        #         writer.writerow(data)
        #
        # except:
        #     with open("ScrapedData.csv","w",encoding='utf-8',newline="") as file:
        #         writer = csv.writer(file)
        #         writer.writerow(["Name","First name","Last name","Title","Email","Phone Number","Linkedin",
        #                          "Employee","Company Website","Company Name","Location","Linkedin Sales Navigator"])
        #         writer.writerow(data)

        try:
            with open("history.txt","a+",encoding='utf-8') as file:
                file.writelines(data[6]+"\n")
        except:
            with open("history.txt","w",encoding='utf-8') as file:
                file.writelines(data[6]+"\n")
        return

# if __name__ == "__main__":
#     obj =  Tool()
#     obj.workbook_name="SalesData.xlsx"
#     obj.sheet_name="SalesNavigator"
#     links=obj.get_linkedIn_data(obj.workbook_name)
#     print(links)
#
#     obj.initialize("testtechhh2999@gmail.com","Crawler12345@")
#     for link in links:
#         # obj.scrape_salesql(link[6])
#         if link[6] in obj.read_history():
#             pass
#         else:
#             if obj.check_url(link[6])==False:
#                 try:
#                     with open("brokenLinks.txt","a+") as file:
#                         file.writelines(link[6]+"\n")
#                 except:
#                     with open("brokenLinks.txt","w") as file:
#                         file.writelines(link[6]+"\n")
#                 link[4]="404 Not Found"
#                 link[5] = "404 Not Found"
#                 link[8] = "404 Not Found"
#                 obj.write_data_v2(link)
#                 continue
#             print("Checking with Apollo")
#             email,number=obj.scrape_apollo(link[6])
#             if email ==None:
#                 try:
#                     # print("Checking with Salesql")
#                     # email=obj.scrape_salesql(link[6])
#                     if email==None:
#                         #call kendo here
#                         print("Checking with Kendo")
#                         email=obj.scrape_kendo(link[6])
#                         if email==None:
#                             print("Checking with Pipileads")
#                             email=obj.scrape_pipileads(link[6])
#                 except:
#                     pass
#             #write data
#             if email==None:
#                 email="NA"
#                 link[8] = "NA"
#             else:
#                 print(email)
#                 try:
#                     if "@gmail.com" in email or "@yahoo.com" in email or "@gmx.com" in email or "@mail.ru" in email:
#                         link[8]="NA"
#                     else:
#                         try:
#                             link[8] = email.split("@")[1]
#                         except:
#                             link[8]="NA"
#                 except:
#                     link[8]="NA"
#             link[4]=email
#             if number==None:
#                 number="NA"
#             else:
#                 if "+" not in number:
#                     number="+"+number
#
#             link[5]=number.replace("-","").replace("/","").replace("\\","").replace(")","").replace("(","").replace(" ","")
#             obj.write_data_v2(link)




